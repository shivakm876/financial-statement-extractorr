"""
Financial Statement Extractor — BULLETPROOF VERSION
Strictly isolates Income Statement section and filters all junk
"""

import os
import re
from io import BytesIO
from typing import List, Any, Dict, Tuple, Optional

import pdfplumber
from pdf2image import convert_from_bytes
import pytesseract

from fastapi import FastAPI, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv

load_dotenv()

# ─────────────────────────────────────────
# Windows Config
# ─────────────────────────────────────────

pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
POPPLER_PATH = r"C:\Users\HP\Downloads\Release-25.12.0-0\poppler-25.12.0\Library\bin"

# ─────────────────────────────────────────
# FastAPI Setup
# ─────────────────────────────────────────

app = FastAPI(title="Financial Statement Extractor", version="11.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─────────────────────────────────────────
# STANDARDIZED SCHEMA (19 line items)
# ─────────────────────────────────────────

SCHEMA_KEYS = [
    "revenue",
    "other_income", 
    "cost_of_materials",
    "purchases_of_stock",
    "employee_expenses",
    "finance_costs",
    "depreciation",
    "other_expenses",
    "total_expenses",
    "profit_before_tax",
    "tax_expense",
    "profit_after_tax",
    "other_comprehensive_income",
    "total_comprehensive_income",
    "ebitda",
    "ebit",
    "gross_profit",
    "operating_profit",
    "net_profit_margin"
]

SCHEMA_LABELS = {
    "revenue": "Revenue",
    "other_income": "Other Income",
    "cost_of_materials": "Cost of Materials",
    "purchases_of_stock": "Purchases of Stock-in-Trade",
    "employee_expenses": "Employee Benefit Expenses",
    "finance_costs": "Finance Costs",
    "depreciation": "Depreciation & Amortization",
    "other_expenses": "Other Expenses",
    "total_expenses": "Total Expenses",
    "profit_before_tax": "Profit Before Tax",
    "tax_expense": "Tax Expense",
    "profit_after_tax": "Profit After Tax (PAT)",
    "other_comprehensive_income": "Other Comprehensive Income",
    "total_comprehensive_income": "Total Comprehensive Income",
    "ebitda": "EBITDA",
    "ebit": "EBIT",
    "gross_profit": "Gross Profit",
    "operating_profit": "Operating Profit",
    "net_profit_margin": "Net Profit Margin (%)"
}

# ─────────────────────────────────────────
# STRICT JUNK FILTERS - THE KEY FIX
# ─────────────────────────────────────────

def is_absolute_junk(label: str) -> bool:
    """Returns True if label is complete garbage that must be filtered"""
    
    if not label or len(label.strip()) < 3:
        return True
    
    label = label.strip()
    
    # Filter pure symbols/formatting
    if re.match(r'^[\s\-_\|\/\\\(\)\[\]\{\}\.,:;]+$', label):
        return True
    
    # Filter single characters or very short
    if len(label) <= 2:
        return True
    
    # Filter parentheses labels like (a), (b), (i), (ii)
    if re.match(r'^\([a-z0-9]+\)$', label, re.I):
        return True
    
    # Filter pure numbers
    if re.match(r'^[\d\s\.,\-]+$', label):
        return True
    
    return False


def is_document_metadata(label: str) -> bool:
    """Returns True if this is document header/footer/metadata"""
    
    label_lower = label.lower().strip()
    
    # Document references
    metadata_patterns = [
        r'^ref:', r'^date:', r'^to:', r'^from:', r'^subject:', r'^sub:',
        r'mumbai', r'delhi', r'bangalore', r'kolkata', r'chennai',
        r'cin:', r'telephone', r'fax', r'email', r'website',
        r'registered office', r'regd\.', r'corporate office',
        r'board of directors', r'chairman', r'director', r'secretary',
        r'din:', r'member', r'auditor',
        r'quarter ended', r'year ended', r'period ended',
        r'unaudited', r'audited', r'reviewed',
        r'compliance', r'regulation', r'sebi', r'listing',
        r'annexure', r'enclosure', r'attachment',
        r'in terms of', r'pursuant to', r'as per', r'based on',
        r'note:', r'notes to', r'see note', r'refer note',
        r'the board', r'the company', r'we have',
        r'place:', r'signed', r'for and on behalf'
    ]
    
    for pattern in metadata_patterns:
        if re.search(pattern, label_lower):
            return True
    
    return False


def is_non_income_statement(label: str) -> bool:
    """Returns True if this is NOT an income statement item"""
    
    label_lower = label.lower().strip()
    
    # Balance sheet items
    balance_sheet_keywords = [
        'equity share capital', 'reserves', 'shareholders', 'equity attributable',
        'non-controlling interest', 'total equity',
        'borrowings', 'lease liabilities', 'deferred tax liabilities',
        'trade payables', 'other financial liabilities', 'provisions',
        'current liabilities', 'non-current liabilities',
        'property plant', 'right-of-use', 'goodwill', 'intangible',
        'investments', 'inventories', 'trade receivables',
        'cash and cash', 'other financial assets', 'current assets',
        'total assets', 'total liabilities'
    ]
    
    # Segment/other sections
    other_sections = [
        'segment revenue', 'segment result', 'segment assets', 'segment liabilities',
        'geographical', 'product segment',
        'earnings per share', 'basic eps', 'diluted eps',
        'paid-up equity', 'face value', 'reserves excluding',
        'debt equity ratio', 'current ratio', 'return on equity'
    ]
    
    all_non_income = balance_sheet_keywords + other_sections
    
    for keyword in all_non_income:
        if keyword in label_lower:
            return True
    
    return False


def is_valid_income_statement_label(label: str) -> bool:
    """Returns True ONLY if this is clearly an income statement line item"""
    
    label_lower = label.lower().strip()
    
    # Must contain at least one of these financial keywords
    income_statement_keywords = [
        'revenue', 'income', 'sales', 'turnover',
        'cost', 'expense', 'expenditure',
        'profit', 'loss', 'surplus', 'deficit',
        'tax', 'depreciation', 'amortization', 'amortisation',
        'finance cost', 'interest', 
        'ebitda', 'ebit', 'pbt', 'pat',
        'operating', 'gross', 'net',
        'employee', 'material', 'purchase',
        'comprehensive'
    ]
    
    has_keyword = any(kw in label_lower for kw in income_statement_keywords)
    
    return has_keyword


# ─────────────────────────────────────────
# MASTER FILTER - COMBINES ALL CHECKS
# ─────────────────────────────────────────

def should_keep_row(label: str) -> bool:
    """Master filter - returns True only if row should be kept"""
    
    # Step 1: Remove absolute junk
    if is_absolute_junk(label):
        return False
    
    # Step 2: Remove metadata
    if is_document_metadata(label):
        return False
    
    # Step 3: Remove non-income statement items
    if is_non_income_statement(label):
        return False
    
    # Step 4: Must be valid income statement label
    if not is_valid_income_statement_label(label):
        return False
    
    return True


# ─────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────

def _clean_cell(cell: Any) -> str:
    if cell is None:
        return ""
    return re.sub(r"\s+", " ", str(cell)).strip()


def _parse_numeric_token(raw: str) -> Optional[float]:
    if not raw:
        return None
    
    token = str(raw).strip()
    
    # Remove common text
    token = re.sub(r'[^\d\.,\-\(\)]', '', token)
    
    if not token or not re.search(r'\d', token):
        return None
    
    # Handle negative numbers in parentheses
    neg = token.startswith("(") and token.endswith(")")
    clean = token.strip("()").replace(",", "")
    
    try:
        val = float(clean)
        return -val if neg else val
    except:
        return None


# ─────────────────────────────────────────
# IMPROVED Table Parser
# ─────────────────────────────────────────

def find_income_statement_table(tables):
    """Find the table that contains the income statement"""
    
    best_table = None
    best_score = 0
    
    for table in tables:
        if not table or len(table) < 8:  # Need minimum rows
            continue
        
        # Score this table
        score = 0
        has_revenue = False
        has_years = False
        
        # Check first few rows for year headers
        for row in table[:3]:
            text = " ".join([_clean_cell(c) for c in row]).lower()
            if re.search(r'(quarter|nine months|year ended|20\d{2}|fy\s?\d{2})', text):
                has_years = True
                score += 10
        
        # Check if table has revenue/income rows
        for row in table:
            label = _clean_cell(row[0]).lower()
            if 'revenue' in label or 'income' in label:
                has_revenue = True
                score += 5
            if 'profit' in label or 'expense' in label:
                score += 2
        
        # Must have both years and revenue to be valid
        if has_years and has_revenue and score > best_score:
            best_score = score
            best_table = table
    
    return best_table


def parse_income_statement_table(table):
    """Parse a validated income statement table"""
    
    rows = []
    years = []
    year_indices = []
    
    # Find header row with years
    header_row_idx = None
    for r_idx, row in enumerate(table[:5]):  # Check first 5 rows
        cells = [_clean_cell(c) for c in row]
        
        temp_indices = []
        temp_years = []
        
        for i, cell in enumerate(cells):
            # Match year patterns
            if re.search(r'(Quarter ended|Preceding quarter|Nine months|Year ended|Corresponding)', cell, re.I):
                temp_indices.append(i)
                temp_years.append(cell.strip())
            elif re.search(r'(31/12/2025|30/09/2025|31/03/2025|31/12/2024)', cell):
                temp_indices.append(i)
                temp_years.append(cell.strip())
            elif re.search(r'(20\d{2}|FY\s?\d{2})', cell) and len(cell.strip()) < 15:
                temp_indices.append(i)
                temp_years.append(cell.strip())
        
        if len(temp_indices) >= 2:  # Need at least 2 year columns
            header_row_idx = r_idx
            year_indices = temp_indices
            years = temp_years
            break
    
    if header_row_idx is None or not year_indices:
        return [], []
    
    # Remove duplicate years (if same year appears twice, keep unique ones)
    unique_years = []
    unique_indices = []
    seen = set()
    for idx, year in zip(year_indices, years):
        # Create a key that's more permissive
        year_key = re.sub(r'\s+', ' ', year.lower().strip())
        if year_key not in seen:
            seen.add(year_key)
            unique_indices.append(idx)
            unique_years.append(year)
    
    year_indices = unique_indices
    years = unique_years
    
    # Extract data rows
    for row in table[header_row_idx + 1:]:
        cells = [_clean_cell(c) for c in row]
        
        if not cells or len(cells) < max(year_indices) + 1:
            continue
        
        label = cells[0]
        
        if not label:
            continue
        
        # Apply master filter
        if not should_keep_row(label):
            continue
        
        # Extract values for each year column
        values = []
        for idx in year_indices:
            if idx < len(cells):
                val = _parse_numeric_token(cells[idx])
                values.append(val)
            else:
                values.append(None)
        
        # Only keep if has at least one number
        if any(v is not None for v in values):
            rows.append({
                "label": label,
                "values": values
            })
    
    return rows, years


# ─────────────────────────────────────────
# PDF EXTRACTION
# ─────────────────────────────────────────

def extract_pdf(pdf_bytes: bytes):
    """Extract financial data from PDF"""
    
    tables = []
    
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            extracted = page.extract_tables()
            if extracted:
                tables.extend(extracted)
    
    # Find best income statement table
    best_table = find_income_statement_table(tables)
    
    if not best_table:
        return {"rows": [], "years": []}
    
    # Parse it
    rows, years = parse_income_statement_table(best_table)
    
    # Validate results
    if len(rows) >= 5 and len(years) >= 2:
        return {"rows": rows, "years": years}
    
    return {"rows": [], "years": []}


# ─────────────────────────────────────────
# METADATA EXTRACTION
# ─────────────────────────────────────────

def extract_metadata(pdf_bytes: bytes) -> Dict[str, str]:
    """Extract company name, currency, units from PDF"""
    
    text = ""
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages[:3]:
            text += page.extract_text() or ""
            if len(text) > 3000:
                break
    
    text = text[:3000]
    
    metadata = {
        "company": "Unknown Company",
        "currency": "INR",
        "unit": "Crores"
    }
    
    # Extract company name
    company_patterns = [
        r"([A-Z][A-Za-z\s&]+(?:Limited|Ltd|Inc|Corporation|Corp))",
    ]
    
    for pattern in company_patterns:
        match = re.search(pattern, text)
        if match:
            company = match.group(1).strip()
            # Filter out common false matches
            if company not in ["Stock Exchange of India Limited", "Listing Department"]:
                metadata["company"] = company
                break
    
    # Extract currency
    if re.search(r"₹|INR|Rupees", text, re.I):
        metadata["currency"] = "INR"
    elif re.search(r"\$|USD|Dollars", text, re.I):
        metadata["currency"] = "USD"
    
    # Extract units
    if re.search(r"Crore|Crores", text, re.I):
        metadata["unit"] = "Crores"
    elif re.search(r"Lakh|Lakhs", text, re.I):
        metadata["unit"] = "Lakhs"
    elif re.search(r"Million|Millions", text, re.I):
        metadata["unit"] = "Millions"
    
    return metadata


# ─────────────────────────────────────────
# INTELLIGENT LABEL MAPPING
# ─────────────────────────────────────────

def map_label_to_schema(label: str) -> Tuple[str, float]:
    """Map extracted label to standardized schema key"""
    
    label_lower = label.lower()
    
    # REVENUE
    if any(x in label_lower for x in ["revenue from operations", "sales", "total revenue", "net sales"]):
        return ("revenue", 1.0)
    if "revenue" in label_lower and "other" not in label_lower:
        return ("revenue", 0.8)
    
    # OTHER INCOME
    if any(x in label_lower for x in ["other income", "non-operating income"]):
        return ("other_income", 1.0)
    
    # COST OF MATERIALS
    if any(x in label_lower for x in ["cost of materials consumed", "cost of goods sold", "cogs"]):
        return ("cost_of_materials", 1.0)
    
    # PURCHASES
    if any(x in label_lower for x in ["purchase of products", "purchases of stock"]):
        return ("purchases_of_stock", 1.0)
    
    # EMPLOYEE EXPENSES
    if any(x in label_lower for x in ["employee benefit", "employee cost", "salaries", "wages"]):
        return ("employee_expenses", 1.0)
    
    # FINANCE COSTS
    if any(x in label_lower for x in ["finance cost", "interest", "borrowing cost"]):
        return ("finance_costs", 1.0)
    
    # DEPRECIATION
    if any(x in label_lower for x in ["depreciation", "amortisation", "amortization"]):
        return ("depreciation", 1.0)
    
    # OTHER EXPENSES
    if any(x in label_lower for x in ["other expense", "administrative", "selling"]):
        return ("other_expenses", 1.0)
    
    # TOTAL EXPENSES
    if any(x in label_lower for x in ["total expense", "total expenditure"]):
        return ("total_expenses", 1.0)
    
    # PROFIT BEFORE TAX
    if any(x in label_lower for x in ["profit before tax", "pbt", "profit before exceptional"]):
        return ("profit_before_tax", 1.0)
    
    # TAX
    if any(x in label_lower for x in ["tax expense", "current tax", "income tax"]) and "profit" not in label_lower:
        return ("tax_expense", 1.0)
    
    # PROFIT AFTER TAX
    if any(x in label_lower for x in ["profit after tax", "pat", "net profit for"]):
        return ("profit_after_tax", 1.0)
    
    # OTHER COMPREHENSIVE INCOME
    if "other comprehensive" in label_lower:
        return ("other_comprehensive_income", 1.0)
    
    # TOTAL COMPREHENSIVE INCOME
    if "total comprehensive" in label_lower:
        return ("total_comprehensive_income", 1.0)
    
    return (None, 0.0)


# ─────────────────────────────────────────
# SCHEMA NORMALIZATION
# ─────────────────────────────────────────

def normalize_to_schema(raw_rows: List[Dict], years: List[str]) -> Dict:
    """Convert raw extracted rows into standardized schema"""
    
    schema_data = {}
    for key in SCHEMA_KEYS:
        schema_data[key] = {
            "values": [None] * len(years),
            "confidence": "low",
            "note": "Not found in document",
            "source": None
        }
    
    for row in raw_rows:
        schema_key, confidence = map_label_to_schema(row["label"])
        
        if schema_key and schema_key in schema_data:
            current_conf = {"high": 1.0, "medium": 0.5, "low": 0.0}[schema_data[schema_key]["confidence"]]
            
            if confidence >= current_conf:
                conf_label = "high" if confidence >= 0.9 else "medium" if confidence >= 0.5 else "low"
                
                schema_data[schema_key] = {
                    "values": row["values"],
                    "confidence": conf_label,
                    "note": f"Mapped from: {row['label']}" if confidence < 1.0 else "Direct match",
                    "source": row["label"]
                }
    
    schema_data = calculate_missing_fields(schema_data, len(years))
    
    return schema_data


# ─────────────────────────────────────────
# CALCULATED FIELDS
# ─────────────────────────────────────────

def calculate_missing_fields(schema_data: Dict, num_years: int) -> Dict:
    """Calculate EBITDA, EBIT, etc."""
    
    def get_vals(key):
        return schema_data.get(key, {}).get("values", [None] * num_years)
    
    # EBITDA = PBT + Finance + Depreciation
    if schema_data["ebitda"]["confidence"] == "low":
        pbt = get_vals("profit_before_tax")
        fin = get_vals("finance_costs")
        dep = get_vals("depreciation")
        
        ebitda_vals = []
        for i in range(num_years):
            if all(v is not None for v in [pbt[i], fin[i], dep[i]]):
                ebitda_vals.append(pbt[i] + fin[i] + dep[i])
            else:
                ebitda_vals.append(None)
        
        if any(v is not None for v in ebitda_vals):
            schema_data["ebitda"] = {
                "values": ebitda_vals,
                "confidence": "medium",
                "note": "Calculated: PBT + Finance Costs + Depreciation",
                "source": "calculated"
            }
    
    # EBIT = EBITDA - Depreciation
    if schema_data["ebit"]["confidence"] == "low":
        ebitda = get_vals("ebitda")
        dep = get_vals("depreciation")
        
        ebit_vals = []
        for i in range(num_years):
            if ebitda[i] is not None and dep[i] is not None:
                ebit_vals.append(ebitda[i] - dep[i])
            else:
                ebit_vals.append(None)
        
        if any(v is not None for v in ebit_vals):
            schema_data["ebit"] = {
                "values": ebit_vals,
                "confidence": "medium",
                "note": "Calculated: EBITDA - Depreciation",
                "source": "calculated"
            }
    
    # Gross Profit
    if schema_data["gross_profit"]["confidence"] == "low":
        rev = get_vals("revenue")
        cogs = get_vals("cost_of_materials")
        purch = get_vals("purchases_of_stock")
        
        gp_vals = []
        for i in range(num_years):
            if rev[i] is not None:
                val = rev[i]
                if cogs[i] is not None:
                    val -= cogs[i]
                if purch[i] is not None:
                    val -= purch[i]
                gp_vals.append(val)
            else:
                gp_vals.append(None)
        
        if any(v is not None for v in gp_vals):
            schema_data["gross_profit"] = {
                "values": gp_vals,
                "confidence": "medium",
                "note": "Calculated: Revenue - COGS - Purchases",
                "source": "calculated"
            }
    
    # Net Profit Margin
    if schema_data["net_profit_margin"]["confidence"] == "low":
        pat = get_vals("profit_after_tax")
        rev = get_vals("revenue")
        
        npm_vals = []
        for i in range(num_years):
            if pat[i] is not None and rev[i] is not None and rev[i] != 0:
                npm_vals.append(round((pat[i] / rev[i]) * 100, 2))
            else:
                npm_vals.append(None)
        
        if any(v is not None for v in npm_vals):
            schema_data["net_profit_margin"] = {
                "values": npm_vals,
                "confidence": "medium",
                "note": "Calculated: (PAT / Revenue) × 100",
                "source": "calculated"
            }
    
    return schema_data


# ─────────────────────────────────────────
# Excel Builder
# ─────────────────────────────────────────

def build_excel(schema_data: Dict, years: List[str], metadata: Dict):
    wb = Workbook()
    ws = wb.active
    ws.title = "Income Statement"
    
    # Title
    ws.append(["Income Statement"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(years) + 2)
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Subtitle
    subtitle = f"{metadata['company']} | Units: {metadata['currency']} {metadata['unit']}"
    ws.append([subtitle])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(years) + 2)
    ws['A2'].font = Font(size=11, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.append([])
    
    # Header
    header = ["Particulars"] + years + ["Analyst Notes"]
    ws.append(header)
    for cell in ws[4]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Colors
    colors = {
        "high": "C6EFCE",
        "medium": "FFEB9C",
        "low": "FFC7CE"
    }
    
    # Data
    for key in SCHEMA_KEYS:
        data = schema_data[key]
        
        formatted_vals = []
        for val in data["values"]:
            if val is None:
                formatted_vals.append("—")
            else:
                formatted_vals.append(round(val, 2))
        
        row = [SCHEMA_LABELS[key]] + formatted_vals + [data["note"]]
        ws.append(row)
        
        row_num = ws.max_row
        fill_color = colors[data["confidence"]]
        
        for col_num in range(2, 2 + len(years)):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    for i in range(len(years)):
        ws.column_dimensions[chr(66 + i)].width = 20
    ws.column_dimensions[chr(66 + len(years))].width = 50
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────
# API Endpoints
# ─────────────────────────────────────────

@app.post("/extract")
async def extract(files: List[UploadFile] = File(...)):
    all_rows = []
    years = []
    metadata = {}

    for f in files:
        pdf_bytes = await f.read()
        
        extracted = extract_pdf(pdf_bytes)
        all_rows.extend(extracted["rows"])
        
        if extracted["years"]:
            years = extracted["years"]
        
        if not metadata:
            metadata = extract_metadata(pdf_bytes)

    if not years:
        years = ["Period 1"]

    schema_data = normalize_to_schema(all_rows, years)
    excel = build_excel(schema_data, years, metadata)

    filename = f"{metadata['company'].replace(' ', '_')}_Income_Statement.xlsx"

    return StreamingResponse(
        excel,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.post("/preview")
async def preview(files: List[UploadFile] = File(...)):
    all_rows = []
    years = []
    metadata = {}

    for f in files:
        pdf_bytes = await f.read()
        extracted = extract_pdf(pdf_bytes)
        all_rows.extend(extracted["rows"])
        
        if extracted["years"]:
            years = extracted["years"]
        
        if not metadata:
            metadata = extract_metadata(pdf_bytes)

    schema_data = normalize_to_schema(all_rows, years)
    
    confidence_summary = {"high": 0, "medium": 0, "low": 0}
    for key in SCHEMA_KEYS:
        conf = schema_data[key]["confidence"]
        confidence_summary[conf] += 1

    return {
        "company": metadata.get("company", "Unknown"),
        "currency": metadata.get("currency", "INR"),
        "unit": metadata.get("unit", "Crores"),
        "years_detected": years,
        "raw_rows_found": len(all_rows),
        "files": len(files),
        "confidence_summary": confidence_summary
    }


@app.get("/")
async def serve_index():
    return FileResponse("index.html") 